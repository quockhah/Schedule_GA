import tkinter as tk
from tkinter import messagebox, filedialog
import random
import datetime
import GeneticAlgorithm
from ScheduleViewer import ScheduleViewer
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

class ExamScheduler:
    def __init__(self, root):
        self.root = root
        self.root.title("Chương trình xếp lịch coi thi cho Giảng Viên Khoa CNTT")
        self.root.geometry("1200x750")

        self.lecturers = []
        self.rooms = []
        self.num_weeks = 1  # Default to 1 week
        self.current_schedule_viewer = None  # Biến lưu trữ lịch hiện tại
        self.days_of_week = ["Thứ Hai", "Thứ Ba", "Thứ Tư", "Thứ Năm", "Thứ Sáu", "Thứ Bảy", "Chủ Nhật"]
        self.create_menu()
        self.create_input_frame()
        self.navigation_frame = tk.Frame(self.root)
        self.navigation_frame.pack(side=tk.TOP, fill=tk.X)
        
        #self.weeks_of_semester = ["Tuần 1", "Tuần 2"] 
        self.export_button = tk.Button(self.navigation_frame, text="Xuất Excel", command=self.export_to_excel)
        self.export_button.pack(side="left", padx=10, pady=5)
        
    def create_menu(self):
        menu_bar = tk.Menu(self.root)
        self.root.config(menu=menu_bar)
        menu_bar.add_command(label="Thoát", command=self.root.quit)

    def create_input_frame(self):
        frame_input = tk.Frame(self.root, relief=tk.GROOVE, bd=2)
        frame_input.pack(pady=10, fill=tk.X)
        # Container for Room and Week Management
        frame_top = tk.Frame(frame_input)
        frame_top.pack(side=tk.TOP, padx=10, pady=5)
        # Lecturer Management
        frame_lecturers = tk.LabelFrame(frame_input, text="Quản lý Giảng viên", font=("Arial", 10))
        frame_lecturers.pack(side=tk.LEFT, padx=10, pady=5)

        self.listbox_lecturers = tk.Listbox(frame_lecturers, height=10, width=30)
        self.listbox_lecturers.pack(side=tk.LEFT, padx=5, pady=5)

        lecturer_buttons = tk.Frame(frame_lecturers)
        lecturer_buttons.pack(side=tk.RIGHT, padx=5)

        tk.Button(lecturer_buttons, text="Thêm", command=self.add_lecturer).pack(fill=tk.X, pady=2)
        tk.Button(lecturer_buttons, text="Xóa", command=self.remove_lecturer).pack(fill=tk.X, pady=2)
        tk.Button(lecturer_buttons, text="Xóa tất cả", command=self.clear_lecturers).pack(fill=tk.X, pady=2)
        tk.Button(lecturer_buttons, text="Lấy từ file", command=self.load_lecturers_from_file).pack(fill=tk.X, pady=2)

        self.entry_lecturer = tk.Entry(frame_lecturers, font=("Arial", 10))
        self.entry_lecturer.pack(fill=tk.X, padx=5, pady=5)
    # Room Management
        frame_rooms = tk.LabelFrame(frame_input, text="Quản lý Phòng thi", font=("Arial", 10))
        frame_rooms.pack(side=tk.LEFT, padx=10, pady=5)
        tk.Label(frame_rooms, text="Số lượng phòng thi:").pack(pady=5)
        self.spinbox_rooms = tk.Spinbox(frame_rooms, from_=1, to=100, width=5, font=("Arial", 12))
        self.spinbox_rooms.pack(pady=5)

        # Week Management
        frame_weeks = tk.LabelFrame(frame_input, text="Quản lý Số Tuần Thi", font=("Arial", 10))
        frame_weeks.pack(side=tk.LEFT, padx=10, pady=5)
        tk.Label(frame_weeks, text="Số tuần thi:").pack(pady=5)
        self.spinbox_weeks = tk.Spinbox(frame_weeks, from_=1, to=10, width=5, font=("Arial", 12))
        self.spinbox_weeks.pack(pady=5)

        # Start Date
        frame_start_date = tk.LabelFrame(frame_input, text="Ngày bắt đầu", font=("Arial", 10))
        frame_start_date.pack(side=tk.LEFT, padx=10, pady=5)
        self.entry_start_date = tk.Entry(frame_start_date, font=("Arial", 10))
        self.entry_start_date.pack(fill=tk.X, padx=5, pady=5)
        # Buttons
        self.btn_confirm = tk.Button(frame_input, text="Xếp Lịch", command=self.confirm_info)
        self.btn_confirm.pack(pady=10)
        
        self.btn_view_saved_schedule = tk.Button(frame_input, text="Xem Lịch Đã Lưu", command=self.view_saved_schedule)
        self.btn_view_saved_schedule.pack(pady=10)


    def confirm_info(self):
        num_rooms = int(self.spinbox_rooms.get())
        self.num_weeks = int(self.spinbox_weeks.get())
        self.rooms = [f"Phòng {i + 1}" for i in range(num_rooms)]

        start_date_str = self.entry_start_date.get().strip()
        try:
            start_date = datetime.datetime.strptime(start_date_str, "%d/%m/%Y")
        except ValueError:
            messagebox.showerror("Lỗi", "Vui lòng nhập ngày bắt đầu hợp lệ (dd/mm/yyyy)!")
            return

        # Kiểm tra ngày bắt đầu không nhỏ hơn ngày hiện tại
        today = datetime.datetime.today()
        if start_date < today:
            messagebox.showerror("Lỗi", "Ngày bắt đầu không được thấp hơn ngày hiện tại!")
            return

        if not self.lecturers or not self.rooms:
            messagebox.showerror("Lỗi", "Vui lòng nhập đủ danh sách giảng viên và phòng thi!")
            return

        # Tạo danh sách ngày nhưng loại bỏ Chủ Nhật
        self.schedule_dates = []
        for i in range(self.num_weeks * 7):
            current_date = start_date + datetime.timedelta(days=i)
            if current_date.weekday() != 6:  # 6 là Chủ Nhật
                self.schedule_dates.append(current_date)

        global NUM_PROCTORS, NUM_ROOMS, DAYS
        NUM_PROCTORS = len(self.lecturers)
        NUM_ROOMS = len(self.rooms)
        DAYS = len(self.schedule_dates)
        genetic_algo = GeneticAlgorithm.GeneticAlgorithm(
            num_proctors=NUM_PROCTORS,
            num_rooms=NUM_ROOMS,
            num_days=len(self.schedule_dates),
            valid_dates=self.schedule_dates,  # Truyền danh sách ngày hợp lệ
            num_slots_per_day=2,  # Giả sử có 2 ca mỗi ngày
            population_size=50,  # Kích thước quần thể
            generations=100,  # Số thế hệ
            mutation_rate=0.1,  # Tỉ lệ đột biến
            proctors_per_room= 3   # Truyền tham số proctors_per_room
        )
        best_schedule = genetic_algo.genetic_algorithm()  # Gọi hàm genetic_algorithm
        self.schedule = best_schedule  # Lưu lịch vào thuộc tính self.schedule

        # Xóa lịch cũ (nếu có)
        if self.current_schedule_viewer is not None:
            self.current_schedule_viewer.day_frame.destroy()
            self.current_schedule_viewer.navigation_frame.destroy()
            self.current_schedule_viewer = None

        # Hiển thị lịch mới
        self.current_schedule_viewer = ScheduleViewer(self.root, best_schedule, self.lecturers, self.schedule_dates, self.rooms)
    
    def add_lecturer(self):
        lecturer_name = self.entry_lecturer.get().strip()
        if lecturer_name and lecturer_name not in self.lecturers:
            self.lecturers.append(lecturer_name)
            self.listbox_lecturers.insert(tk.END, lecturer_name)
            self.entry_lecturer.delete(0, tk.END)
        else:
            messagebox.showwarning("Cảnh báo", "Tên giảng viên không hợp lệ hoặc đã tồn tại!")

    def remove_lecturer(self):
        selected = self.listbox_lecturers.curselection()
        if selected:
            lecturer_name = self.listbox_lecturers.get(selected)
            self.lecturers.remove(lecturer_name)
            self.listbox_lecturers.delete(selected)

    def clear_lecturers(self):
        self.lecturers.clear()
        self.listbox_lecturers.delete(0, tk.END)

    def load_lecturers_from_file(self):
        file_path = filedialog.askopenfilename(filetypes=[("Text Files", "*.txt")])
        if not file_path:
            return

        try:
            with open(file_path, "r", encoding="utf-8") as file:
                lines = file.readlines()
                new_lecturers = [line.strip() for line in lines if line.strip()]

                for lecturer in new_lecturers:
                    if lecturer not in self.lecturers:
                        self.lecturers.append(lecturer)
                        self.listbox_lecturers.insert(tk.END, lecturer)
        except Exception as e:
            messagebox.showerror("Lỗi", f"Không thể đọc file: {e}")

    # Hàm mở và xem lịch đã lưu
    def view_saved_schedule(self):
        file_path = filedialog.askopenfilename(defaultextension=".txt", filetypes=[("Text Files", "*.txt")])
        if not file_path:
            return

        try:
            with open(file_path, "r", encoding="utf-8") as file:
                saved_schedule = file.read().splitlines()  # Chia nội dung thành các dòng

            # Bỏ qua các dòng chứa thông tin ngày (giả sử dòng chứa "Ngày" là dòng ngày)
            saved_schedule = [line for line in saved_schedule if "Ngày" not in line]

            # Tìm và lưu thông tin ngày bắt đầu và ngày kết thúc
            start_end_dates = [line for line in saved_schedule if "Lịch thi từ ngày" in line]
            
            # Lấy thông tin ngày bắt đầu và kết thúc chỉ một lần
            if start_end_dates:
                date_info = start_end_dates[0].replace("Lịch thi từ ngày", "").strip()
                start_date_str, end_date_str = date_info.split("đến")
                start_date_str = start_date_str.strip()
                end_date_str = end_date_str.strip()

                # Xóa dòng chứa thông tin ngày bắt đầu và kết thúc sau khi đã lấy được thông tin
                saved_schedule = [line for line in saved_schedule if "Lịch thi từ ngày" not in line]
            else:
                start_date_str = "Không có thông tin ngày bắt đầu"
                end_date_str = "Không có thông tin ngày kết thúc"

            # Chuyển start_date_str thành đối tượng datetime
            start_date = datetime.datetime.strptime(start_date_str, "%d/%m/%Y")

            # Mở cửa sổ mới để hiển thị lịch
            saved_schedule_window = tk.Toplevel(self.root)
            saved_schedule_window.title("Lịch Đã Lưu")
            saved_schedule_window.geometry("800x600")

            # Tiêu đề cửa sổ
            tk.Label(saved_schedule_window, text="Lịch thi đã lưu", font=("Arial", 16, "bold")).pack(pady=10)

            # Hiển thị ngày bắt đầu và kết thúc
            date_label = tk.Label(saved_schedule_window, text=f"Ngày: {start_date_str} đến {end_date_str}", 
                                font=("Arial", 14), anchor="center")
            date_label.pack(pady=10)

            # Tạo bảng hiển thị lịch
            table_frame = tk.Frame(saved_schedule_window)
            table_frame.pack(fill="both", expand=True, padx=20, pady=10)

            # Tiêu đề cột (Phòng thi)
            tk.Label(table_frame, text="Ca/Phòng", font=("Arial", 12, "bold"), anchor="center", width=12).grid(row=0, column=0, padx=5, pady=5)
            for room in range(10):
                tk.Label(table_frame, text=f"Phòng {room + 1}", font=("Arial", 12, "bold"), anchor="center", width=12).grid(row=0, column=room + 1, padx=5, pady=5)

            # Nội dung bảng (Các ca thi)
            time_slots = ["Sáng", "Trưa", "Chiều"]  # Các ca thay thế
            row_index = 1

            def display_day_schedule(day_index):
                nonlocal row_index
                # Xóa tất cả các label cũ trong bảng
                for widget in table_frame.winfo_children():
                    widget.grid_forget()

                # Tính toán ngày cho day_index
                current_day = start_date + datetime.timedelta(days=day_index)
                current_day_str = current_day.strftime("%d/%m/%Y")  # Chuyển ngày sang định dạng chuỗi

                # Hiển thị ngày tháng cho ngày hiện tại
                tk.Label(table_frame, text=f"Ngày {current_day_str}", font=("Arial", 14, "bold"), anchor="center", width=12).grid(row=0, columnspan=11, padx=5, pady=5)

                # Hiển thị lại tiêu đề cột (Phòng thi)
                tk.Label(table_frame, text="Ca/Phòng", font=("Arial", 12, "bold"), anchor="center", width=12).grid(row=1, column=0, padx=5, pady=5)
                for room in range(10):
                    tk.Label(table_frame, text=f"Phòng {room + 1}", font=("Arial", 12, "bold"), anchor="center", width=12).grid(row=1, column=room + 1, padx=5, pady=5)

                row_index = 2
                # Duyệt qua lịch đã lưu và hiển thị theo từng ca, từng phòng cho ngày hiện tại
                for slot_index, slot in enumerate(saved_schedule[day_index * 3:(day_index + 1) * 3]):
                    # Tạo label cho từng ca
                    tk.Label(table_frame, text=f"{time_slots[slot_index % len(time_slots)]}", font=("Arial", 12), anchor="center", width=12).grid(row=row_index, column=0, padx=5, pady=5)

                    # Hiển thị tên giảng viên theo từng phòng
                    lecturers = slot.split(',')  # Giả sử danh sách giảng viên trong mỗi ca cách nhau bằng dấu ","
                    for room_index, lecturer_name in enumerate(lecturers):
                        lecturer_label = tk.Label(
                            table_frame,
                            text=lecturer_name.strip(),  # Xóa khoảng trắng thừa
                            font=("Arial", 10),
                            anchor="center",
                            width=12,
                            height=2,
                            wraplength=100
                        )
                        lecturer_label.grid(row=row_index, column=room_index + 1, padx=5, pady=5)

                    row_index += 1

            # Display the first day by default
            display_day_schedule(0)

            # Add buttons for previous and next day
            def prev_day():
                nonlocal current_day_index
                if current_day_index > 0:
                    current_day_index -= 1
                    display_day_schedule(current_day_index)

            def next_day():
                nonlocal current_day_index
                if current_day_index < len(saved_schedule) // 3 - 1:
                    current_day_index += 1
                    display_day_schedule(current_day_index)

            current_day_index = 0

            prev_button = tk.Button(saved_schedule_window, text="Prev Day", command=prev_day)
            prev_button.pack(side="left", padx=10, pady=10)

            next_button = tk.Button(saved_schedule_window, text="Next Day", command=next_day)
            next_button.pack(side="right", padx=10, pady=10)

        except Exception as e:
            messagebox.showerror("Lỗi", f"Không thể mở lịch đã lưu: {e}")

   
    
    def export_to_excel(self, file_name="Exam_Schedule.xlsx"):
        # Tạo một workbook mới và sheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Lịch Thi"

        # Định dạng tiêu đề
        title_font = Font(bold=True, size=14)
        header_font = Font(bold=True)
        center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Ngày bắt đầu và kết thúc
        start_date = self.schedule_dates[0]
        end_date = self.schedule_dates[-1]
        ws["A1"] = f"Lịch thi từ ngày {start_date.strftime('%d/%m/%Y')} đến ngày {end_date.strftime('%d/%m/%Y')}"
        ws["A1"].font = title_font
        ws["A1"].alignment = center_alignment
        ws.merge_cells("A1:J1")

        # Dòng tiêu đề (gộp ô)
        ws.merge_cells("A2:A4")
        ws["A2"] = "STT"
        ws["A2"].alignment = center_alignment
        ws["A2"].font = header_font

        ws.merge_cells("B2:B4")
        ws["B2"] = "Giảng Viên"
        ws["B2"].alignment = center_alignment
        ws["B2"].font = header_font

        # Các cột tiêu đề cho các ngày trong tuần
        columns = ["Sáng", "Chiều"]

        # Tính số cột tối đa dựa trên số ngày
        max_excel_columns = len(self.schedule_dates) * 3

        for day_index, day_date in enumerate(self.schedule_dates):
            day_name = self.days_of_week[day_date.weekday()]
            date_str = day_date.strftime('%d/%m/%Y')
            start_col = 3 + day_index * 3

            # Hợp nhất ô cho ngày
            ws.merge_cells(start_row=2, start_column=start_col, end_row=2, end_column=start_col + 2)
            cell = ws.cell(row=2, column=start_col, value=f"{day_name}, {date_str}")
            cell.alignment = center_alignment
            cell.font = header_font

            # Các buổi trong ngày
            for i, period in enumerate(columns):
                col_index = start_col + i
                cell = ws.cell(row=3, column=col_index, value=period)
                cell.alignment = center_alignment
                cell.font = header_font

        # Điền dữ liệu giảng viên và lịch coi thi
        row = 5
        for idx, lecturer in enumerate(self.lecturers, start=1):
            # Số thứ tự
            ws.cell(row=row, column=1, value=idx).alignment = center_alignment
            
            # Tên giảng viên
            ws.cell(row=row, column=2, value=lecturer).alignment = center_alignment
            
            # Lịch coi thi
            if idx-1 < len(self.schedule):
                for day_index, day_schedule in enumerate(self.schedule[idx-1][:max_excel_columns//3]):
                    for period_index, period in enumerate(day_schedule):
                        if period == 1:
                            col_index = 3 + day_index * 3 + period_index
                            ws.cell(row=row, column=col_index, value="x").alignment = center_alignment
            row += 1

        # Căn giữa tất cả các ô
        for row_cells in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row_cells:
                cell.alignment = center_alignment

        # Điều chỉnh độ rộng cột
        ws.column_dimensions['A'].width = 5   # Cột STT
        ws.column_dimensions['B'].width = 20  # Cột Tên Giảng Viên
        for col in 'CDEFGHIJ':
            ws.column_dimensions[col].width = 10  # Các cột ngày và buổi

        # Lưu file Excel
        wb.save(file_name)
        messagebox.showinfo("Xuất Excel", f"Lịch thi đã được lưu vào file '{file_name}'")
        print(f"Lịch thi đã được lưu vào file '{file_name}'")

